import sqlite3
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# === 1. Leer desde la base ===
conn = sqlite3.connect("historial_precios.db")
df = pd.read_sql_query("SELECT * FROM precios", conn)
conn.close()

# Asegurar que los valores estén en formato numérico
columnas_importe = [
    "Precio 9 de Julio",
    "Precio de lista 9dj",
    "Precio Trenque",
    "Precio de LIsta TL"
]

columnas_porcentaje = [
    "Variacion 9dj",
    "Variacion TL"
]

for col in columnas_importe + columnas_porcentaje:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

# === 2. Calcular las variaciones con la fórmula C/B - 1 ===
df["Variacion 9dj"] = df["Precio de lista 9dj"] / df["Precio 9 de Julio"] - 1
df["Variacion TL"] = df["Precio de LIsta TL"] / df["Precio Trenque"] - 1

# === 3. Exportar sin formato ===
ruta_salida = "exporte_historico.xlsx"
df.to_excel(ruta_salida, index=False)

# === 4. Formatear solo los importes ===
wb = load_workbook(ruta_salida)
ws = wb.active

columnas = {cell.value: idx for idx, cell in enumerate(ws[1])}

# 💲 Formato numérico para importes
for row in ws.iter_rows(min_row=2):
    for col_name in columnas_importe:
        if col_name in columnas:
            celda = row[columnas[col_name]]
            try:
                float(celda.value)
                celda.number_format = '#.##0,00'
            except:
                pass

# 📉 Las variaciones se exportan sin formato, como número decimal

wb.save(ruta_salida)
print(f"✅ Archivo exportado correctamente: {ruta_salida}")
